# 사업자등록상태 조회 자동화
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

browser = webdriver.Chrome()
browser.get(
    "https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/ab/a/a/UTEABAAA13.xml"
)
browser.maximize_window()
time.sleep(1)

# 엑셀작업
wb = load_workbook("./rpabasic/crawl/download/business_number.xlsx")
ws = wb.active
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 100
ws.column_dimensions["B"].width = 11
# 엑셀 전체 내용 출력
# for x in range(2, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(x, y).value, end=" ")
#     print()

for x in range(2, ws.max_row + 1):
    # 사업자등록번호 읽어오기
    bsn = ws.cell(x, 1).value
    browser.find_element(By.ID, "bsno").send_keys(bsn)

    print("==================== 사업자등록번호 : {}".format(bsn))

    # 조회하기 클릭
    browser.find_element(By.ID, "trigger5").click()

    time.sleep(1)

    # 상태 화면 출력
    # tbody = browser.find_element(By.XPATH, '//*[@id="grid2_body_tbody"]')
    # print(tbody.text) # 767-82-00017 부가가치세 일반과세자 입니다. 2022-06-23

    soup = BeautifulSoup(browser.page_source, "lxml")
    tds = soup.select("#grid2_body_tbody > tr > td")

    # 하나의 리스트로 생성
    list1 = []
    for td in tds:
        list1.append(td.get_text())

    # 생성된 리스트 엑셀에 추가
    ws.append(list1)
    time.sleep(0.5)
    del soup

# 기존 내용 삭제
# 2번행부터 10개 삭제
ws.delete_rows(2, 10)

# 저장
wb.save("./rpabasic/crawl/download/business_number_edited.xlsx")
wb.close()

time.sleep(3)
browser.quit()
