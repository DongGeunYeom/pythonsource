# gmarket best - 컴퓨터/전자 항목 추출
from ctypes import alignment
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment

url = "http://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G06"

res = requests.get(url)
soup = BeautifulSoup(res.text, "lxml")

# 1차 카테고리 추출하기
test1 = soup.select("div.best-list li")

# 엑셀 파일 생성
wb = Workbook()

# 기본 시트 활성화
ws = wb.active

# A 컬럼 width 조절
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 80

# 시트명 새로 지정
ws.title = "컴퓨터전자100"

# 제목행 지정
ws.append(["제품 순위", "상품명", "상품가격", "회사명", "상품 URL"])

for idx, item in enumerate(test1):

    # 상품명
    name = item.select_one("a.itemname")
    name2 = name.get_text().lstrip("[").split("]").pop()

    # 가격
    price = item.select_one("div.s-price strong span").get_text()

    # name["href"]를 이용해서 requests.get() 요청
    product_url = name["href"]
    res = requests.get(product_url)
    detail_soup = BeautifulSoup(res.text, "lxml")

    # 회사명 추출(회사명 없으면 셀러명 추출)
    company = detail_soup.select_one("span.text__brand > span.text")

    if not company:
        company = detail_soup.select_one("span.text__seller > a")
    # AttributeError: 'NoneType' object has no attribute 'get_text'

    # 순위, 상품명, 가격, 회사명, 상품상세정보 url

    ws.append([idx + 1, name2, price, company.get_text(), product_url])
    # 하이퍼링크 걸기(제목행 제외)
    ws.cell(row=idx + 2, column=5).hyperlink = product_url

#  파일명 gmarket_오늘날짜.xlsx

# 파일명 : clien_220620.xlsx
today = datetime.now().strftime("%y%m%d")
filename = f"gmarket_{today}.xlsx"

# 서식 지정
font = Font(name="Tahoma", size=14, color="01579b")
alignment = Alignment(horizontal="center")

cell_a1 = ws["A1"]
cell_a1.alignment = alignment
cell_a1.font = font

cell_b1 = ws["B1"]
cell_b1.alignment = alignment
cell_b1.font = font

cell_c1 = ws["C1"]
cell_c1.alignment = alignment
cell_c1.font = font

cell_d1 = ws["D1"]
cell_d1.alignment = alignment
cell_d1.font = font

cell_e1 = ws["E1"]
cell_e1.alignment = alignment
cell_e1.font = font

# 엑셀 저장
wb.save("./rpabasic/crawl/download/" + filename)
