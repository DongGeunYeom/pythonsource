# train.xlsx 읽어서 정규식 적용
# 전체 데이터를 읽어서 성별에 따른 고객명단 이동
# 새로운 엑셀 파일 작성 - 4개 시트 생성
# 남성(Mr.), 미혼여성(Miss.), 기혼여성(Mrs.), 기타

# 5번째 시트 생성(보고서)
# 분류 생존자수 사망자수 생존률
# 남성   48     12     80%
# 미혼여성 ##
# 기혼여성 ##
# 기타   ##

from openpyxl import Workbook
from openpyxl import load_workbook
import re
from requests import patch

path = "./rpabasic/crawl/download/"

# 엑셀 파일 읽기
wb = load_workbook(path + "train.xlsx")
ws = wb.active

# 1) 엑셀 파일 쓰기
work_book = Workbook()

# 2) 시트 생성
work_sheet_man = work_book.create_sheet()
work_sheet_man.title = "남성"
work_sheet_man.column_dimensions["D"].width = 70

work_sheet_women = work_book.create_sheet()
work_sheet_women.title = "미혼여성"
work_sheet_women.column_dimensions["D"].width = 70

work_sheet_married_women = work_book.create_sheet()
work_sheet_married_women.title = "기혼여성"
work_sheet_married_women.column_dimensions["D"].width = 70

work_sheet_others = work_book.create_sheet(title="기타")
# work_sheet_others.title = "기타"
work_sheet_others.column_dimensions["D"].width = 70

# 리포트 시트 생성
work_sheet_report = work_book.create_sheet(title="보고서")
# work_sheet_others.title = "기타"
work_sheet_report.append(["분류", "생존자수", "사망자수", "생존률"])

# 생존자 수, 사망자 수 카운트 할 변수 선언
man_survived, man_unsurvived = 0, 0
solo_survived, solo_unsurvived = 0, 0
married_survived, married_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0

# 엑셀 데이터 다루기
pattern = re.compile(" [A-Za-z]+\.")
for x in ws.iter_rows():

    # 첫 번째 행인 경우 제목행이기 때문에 모든 sheet에 붙여넣기
    # title_list = []
    if x[0].row == 1:
        # for col in x:
        #     title_list.append(col.value)
        # work_sheet_man.append(title_list)
        # work_sheet_married_women.append(title_list)
        # work_sheet_women.append(title_list)
        # work_sheet_others.append(title_list)
        work_sheet_man.append(col.value for col in x)
        work_sheet_married_women.append(col.value for col in x)
        work_sheet_women.append(col.value for col in x)
        work_sheet_others.append(col.value for col in x)

    # 두 번째 ~ 마지막 행까지는 이름 읽어서 각 성별에 맞춰서 시트에 붙여넣기
    else:
        data = pattern.findall(x[3].value)
        survive_data = x[1].value

        if len(data) > 0:
            if data[0] == " Mr.":
                work_sheet_man.append([col.value for col in x])
                if survive_data == 1:  # 생존자
                    man_survived += 1
                else:
                    man_unsurvived += 1

            elif data[0] == " Miss.":
                work_sheet_women.append([col.value for col in x])
                if survive_data == 1:  # 생존자
                    solo_survived += 1
                else:
                    solo_unsurvived += 1

            elif data[0] == " Mrs.":
                work_sheet_married_women.append([col.value for col in x])
                if survive_data == 1:  # 생존자
                    married_survived += 1
                else:
                    married_unsurvived += 1

            else:
                work_sheet_others.append([col.value for col in x])
                if survive_data == 1:  # 생존자
                    others_survived += 1
                else:
                    others_unsurvived += 1


# 보고서 작성
# 남성 생존율
man_survived_rate = "%.2f%%" % (man_survived / (man_survived + man_unsurvived) * 100)

# 미혼여성 생존율
solo_survived_rate = "%.2f%%" % (
    solo_survived / (solo_survived + solo_unsurvived) * 100
)

# 기혼여성 생존율
married_survived_rate = "%.2f%%" % (
    married_survived / (married_survived + married_unsurvived) * 100
)

# 기타 생존율
others_survived_rate = "%.2f%%" % (
    others_survived / (others_survived + others_unsurvived) * 100
)

work_sheet_report.append(["남성", man_survived, man_unsurvived, man_survived_rate])
work_sheet_report.append(["미혼여성", solo_survived, solo_unsurvived, solo_survived_rate])
work_sheet_report.append(
    ["기혼여성", married_survived, married_unsurvived, married_survived_rate]
)
work_sheet_report.append(
    ["기타", others_survived, others_unsurvived, others_survived_rate]
)


work_book.save(path + "train_gender.xlsx")

# 4) 닫기
wb.close()
work_book.close()


# if len(pattern2.findall(x[3].value)) > 0:
#     work_sheet_man.append([x[3].value])

#  print(pattern.search(x[3].value).group())

# pattern3 = re.compile(" Miss\.")
# for x in ws.rows:
#     if len(pattern3.findall(x[3].value)) > 0:
#         work_sheet_women.append([x[3].value])

#      print(pattern.search(x[3].value).group())

# pattern4 = re.compile(" Mrs\.")
# for x in ws.rows:
#     if len(pattern4.findall(x[3].value)) > 0:
#         work_sheet_married_women.append([x[3].value])

# pattern5 = re.compile(" [^Mrs\.]")
# for x in ws.rows:
#     if (
#         len(
#             pattern2.findall(x[3].value)
#             or pattern3.findall(x[3].value)
#             or pattern4.findall(x[3].value)
#         )
#         == 0
#     ):
#         work_sheet_others.append([x[3].value])

# print(pattern.search(x[3].value).group())
