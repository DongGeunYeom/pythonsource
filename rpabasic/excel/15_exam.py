# Quiz) 여러분은 솔대학의 컴퓨터과 교수님....


# - 출석 10
# - 퀴즈1 10
# - 퀴즈2 10
# - 중간고사 20
# - 기말고사 30
# - 프로젝트 20
# _-----------------
# - 총합 100

# 1. 퀴즈 2 점수를 10으로 수정
# 2. H열에 총점(Sum 이용), I열에 성적 정보 추가
# - 총점 90이상 A, 80이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생인 총점 상관없이 F

#  최종 파일명 : scores.xlsx

from json import load
import random
from openpyxl import Workbook
from openpyxl import load_workbook
import win32com.client


def create_table():
    wb = Workbook()
    ws = wb.active

    # 타이틀
    ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "성적정보"])

    # for idx, cell in enumerate(ws["D"]):
    #     if idx == 0 : # 제목행인 경우 skip
    #         continue
    #     cell.value = 10

    # 점수 생성
    for i in range(1, 10):
        # 번호 부여
        ws.cell(row=i + 1, column=1, value=i)
        for j in range(1, 8):
            if j == 2:
                ws.cell(row=i + 1, column=j, value=random.randint(0, 10))
            elif j == 3:
                ws.cell(row=i + 1, column=j, value=random.randint(0, 10))
            elif j == 4:
                # 퀴즈 2 점수 10점으로 변경
                ws.cell(row=i + 1, column=j, value=10)
            elif j == 5:
                ws.cell(row=i + 1, column=j, value=random.randint(0, 20))
            elif j == 6:
                ws.cell(row=i + 1, column=j, value=random.randint(0, 30))
            elif j == 7:
                ws.cell(row=i + 1, column=j, value=random.randint(0, 20))

    # 합계 생성
    for a in range(2, 11):
        ws.cell(row=a, column=8).value = "=sum(B{0}:g{0})".format(a)

    wb.save("./rpabasic/excel/scores.xlsx")


def set_grade():
    excel = win32com.client.Dispatch("Excel.Application")
    temp_wb = excel.Workbooks.Open("./rpabasic/excel/scores.xlsx")

    temp_wb.Save()
    excel.quit()
    wb = load_workbook("./rpabasic/excel/scores.xlsx", data_only=True)
    ws = wb.active

    print(ws["H2"].value)


set_grade()

# 총점, 성적 셀 추가
# for idx, score in enumerate(scores, start=2):
#     ws.cell(row=idx, column=8).value = "=sum(B{0}:g{0})".format(idx)

#     total = sum(score[1:6]) - score[3] + 10
#     grade = None

#     if total >= 90:
#         grade = "A"
#     elif total >= 80:
#         grade = "B"
#     elif total >= 70:
#         grade = "C"
#     else:
#         grade = "D"

#     if score[1] < 5:
#         grade = "F"

#     ws.cell(row=idx, column=9, value=grade)
