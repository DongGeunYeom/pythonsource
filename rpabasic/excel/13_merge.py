# 셀 병합
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2")
ws["B2"].value = "Merged cell"

wb.save("./rpabasic/excel/merge.xlsx")
